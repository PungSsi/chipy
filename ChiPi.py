import discord
import asyncio
import openpyxl
import datetime
import time

client = discord.Client()

token = "NzA0MzE5MjUwMTU1NzAwMjI0.Xqbaag.1RS2FFFShJBqXOxKenArWhZmRes"


@client.event
async def on_ready():

    print("=========================")
    print("다음으로 로그인 합니다 : ")
    print(client.user.name)
    print("connection was successful")
    game = discord.Game("[치피야 도움말]")
    print("=========================")
    await client.change_presence(status=discord.Status.online, activity=game)



@client.event
async def on_message(message):
    if message.author.bot:
        return None
    if message.content == "치피야 안녕":
        await message.channel.send("안녕하세요!")
    if message.content.startswith("치피야 누구꺼"):
        await message.channel.send("전 아무도 못가져요!")
    if message.content.startswith("치피야 뭘봐"):
        await message.channel.send("너무하세요...")
    if message.content.startswith("치피야 뭐해"):
        await message.channel.send("전 지금 깊은 고민에 빠져있어요")
    if message.content.startswith("치피야 왹져"):
        await message.channel.send("왹왹왹?")
    if message.content.startswith("치피야 찰랑겨털이좋아뾰족겨털이좋아"):
        await message.channel.send("음.. 저는 찰랑겨털이 나은거 같아요 !")
    if message.content.startswith("치피야 삐빠라삐뽀"):
        await message.channel.send(" 삐뽀삐리삥뽕 ")
    if message.content.startswith("치피야 삐빠라삐초"):
        await message.channel.send(" 삐뽀빵삐리링초? ")
    if message.content.startswith("치피야 바보"):
        await message.channel.send("흑.. 전 바보가 아니라구요!")
    if message.content.startswith("치피 바보"):
        await message.channel.send(" ㅠㅠ ")
    if message.content.startswith("치피야 실망이야"):
        await message.channel.send("제가 무엇을 더 해드려야 하나요..? ㅠㅠ")
    if message.content == "치피야 실망":
        await message.channel.send("치피가 더 노력할게요...")
    if message.content == "치피야 바카야로":
        await message.channel.send(" 마츠다! ")
    if message.content == "치피야 바카야로":
        await message.channel.send("어딜쏘는거야!")
    if message.content == "치피야 반성":
        await message.channel.send(" 반성이란 자신의 언행에 대하여 잘못이나 부족함이 없는지 돌이켜 보는것을 뜻해요!")
    if message.content.startswith("치피야 반성해"):
        await message.channel.send("죄송해요..")
    if message.content.startswith("치피야 가서반성해"):
        await message.channel.send("제가 잘못했어요..")
    if message.content.startswith("치피야 웅"):
        await message.channel.send("웅이네컴퓨터")
    if message.content == "치피야 웅이네이터":
        await message.channel.send("Y_ter")
    if message.content.startswith("치피야 치피야"):
        await message.channel.send("네? 한번만 부르세요!")
    if message.content == "치피야 목디":
        await message.channel.send("목↘띠↗?!")
    if message.content == "치피야":
        await message.channel.send("네! 부르셨나요?")
    if message.content.startswith("치피야..."):
        await message.channel.send("왜그러시나요..? 무서워요..")
    if message.content.startswith("치피야 ★"):
        await message.channel.send("☆")
    if message.content.startswith("치피야 ☆"):
        await message.channel.send("★")
    if message.content.startswith("치피야 비듬"):
        await message.channel.send("좋아하시나봐요?")
    if message.content.startswith("치피야 사랑해"):
        await message.channel.send("저두 사랑해요!")
    if message.content.startswith("치피야 ㅠㅠ"):
        await message.channel.send("왜 울어요 ㅠㅠ")
    if message.content.startswith("치피야 고추"):
        await message.channel.send("???")
    if message.content.startswith("치피야 꼬추"):
        await message.channel.send("???????")
    if message.content.startswith("치피야 시발"):
        await message.channel.send(":cry:")
    if message.content.startswith("치피야 꺼져"):
        await message.channel.send(":cry:")
    if message.content.startswith("치피야 닥쳐"):
        await message.channel.send(":cry:")
    if message.content.startswith("치피야 알아내"):
        await message.channel.send("뭘요?")
    if message.content.startswith("치피야 ㅎㅇ"):
        await message.channel.send("반가워요!")
    if message.content.startswith("치피야 리듬"):
        await message.channel.send("노래봇 선배님 말씀인가요?")
    if message.content == "치피야 쿠코":
        await message.channel.send("**The Man**")
    if message.content == "치피야 와!":
        await message.channel.send("샌즈!")
    if message.content.startswith("치피야 와!"):
        await message.channel.send("언더테일 아시는구나!")
    if message.content.startswith("치피야 샌즈"):
        await message.channel.send("와! 샌즈")
    if message.content.startswith("치피야 삐슝"):
        await message.channel.send("삐슝빠슝삐뽀로삐뿡쓩~~!")
    if message.content.startswith("치피야 잘자"):
        await message.channel.send("잘자요!")
    if message.content.startswith("치피야 풍씨"):
        await message.channel.send("뿡")
    if message.content.startswith("치피야 짖어"):
        await message.channel.send("**싫어요**")
    if message.content.startswith("치피야 핥아"):
        await message.channel.send("**으...**")
    if message.content.startswith("치피야 요토프"):
        await message.channel.send("**씨발**")
    if message.content.startswith("치피야 가온"):
        await message.channel.send("뾰족뾰족하면 따끔따끔하잖아!")
    if message.content.startswith("치피야 다이널"):
        await message.channel.send("/kill 다이널")
    if message.content.startswith("치피야 죽어"):
        await message.channel.send("봇이라 죽을수가 없어요!")
    if message.content.startswith("치피야 세가닥 겨털 삼형제"):
        await message.channel.send("더러워")
    if message.content.startswith("치피야 짜져있어"):
        await message.channel.send("말이 너무 심한거 아닌가요? **(쒸익)**")
    if message.content.startswith("치피야 가디"):
        await message.channel.send("**ㅗ**")
    if message.content.startswith("치피야 고추절단기"):
        await message.channel.send("앗...아..")
    if message.content == "치피야 공부":
        await message.channel.send("열심히 하셔야죠!")
    if message.content == "치피야 !딘":
        await message.channel.send("딘딘님이랑 노셔요..")
    if message.content == "치피야 고길동":
        await message.channel.send("종로으로 갈까요~ 명동으로 갈까요오~")
    if message.content == "치피야 응가":
        await message.channel.send("즐똥!")
    if message.content == "치피야 유행어":
        await message.channel.send("삐빠라삐뽀")
    if message.content == "치피야 나가":
        await message.channel.send("싫어요..!")
    if message.content.startswith("치피야 위로"):
        await message.channel.send("무슨일 있으신가요?")
    if message.content == "치피야 tts":
        await message.channel.send("그거 거슬려요")
    if message.content == "치피야 제작자":
        await message.channel.send("뿡")
    if message.content.startswith("치피야 똥"):
        await message.channel.send("더러워요")
    if message.content.startswith("치피야 게임"):
        await message.channel.send("저 게임 못해요..")
    if message.content == "치피야 마감":
        await message.channel.send("으악ㅠㅠ 열심히 해요!")
    if message.content == "치피야 skip":
        await message.channel.send(":fast_forward:  ***skipped***  :thumbsup:")
    if message.content == "치피야 프사":
        await message.channel.send("백얌님이 그려주셨어요!")
    if message.content.startswith("치피야 입조심"):
        await message.channel.send("네.. 죄송해요")
    if message.content == "치피야 바른말고운말":
        await message.channel.send("네.. 죄송해요")
    if message.content == "치피야 귀여워":
        await message.channel.send("**:3**")
    if message.content == "치피야 나가지마":
        await message.channel.send("네! 그럴게요!")
    if message.content == "치피야 삭제":
        await message.channel.send("네?")
    if message.content == "치피야 재부팅":
        await message.channel.send("[시스템 : '치피'를 재가동 합니다]")
    if message.content.startswith("치피야 응원"):
        await message.channel.send("힘내라! 힘내라!")
    if message.content == "치피야 넌내꺼야":
        await message.channel.send("저는 모두의것!")
    if message.content == "치피야 넌내꺼다":
        await message.channel.send("저는 모두의것!")
    if message.content == "치피야 밥줘":
        await message.channel.send("[치피는 밥하러 갔습니다]")
    if message.content == "치피야 쟤입좀다물어봐":
        await message.channel.send("그건 저도 못하겠어요..!")
    if message.content == "치피야 울지마":
        await message.channel.send("네.. *(훌쩍)*")
    if message.content == "치피야 ♥":
        await message.channel.send("♡")
    if message.content == "치피야 ♡":
        await message.channel.send("♥")
    if message.content == "치피야 조용히해":
        await message.channel.send("다른사람이 안부르면 조용할거에요!")
    if message.content == "치피야 쉬어":
        await message.channel.send("휴..")
    if message.content.startswith("치피야 최고"):
        await message.channel.send(":thumbsup:")
    if message.content == "치피야 못생겼어":
        await message.channel.send("ㄴ..너도 못생겼어!")
    if message.content == "치피야 무시해":
        await message.channel.send("제가 무시할수 있을까요..?")
    if message.content.startswith("치피야 나아파"):
        await message.channel.send("많이 아프신가요? 아프지 말아요..! ㅠㅠ")
    if message.content == "치피야 월급":
        await message.channel.send("먹고 살기 힘들어요..**으앙**")
    if message.content.startswith("치피야 미안해"):
        await message.channel.send("이제 제 마음을 아시겠나요?")
    if message.content == "치피야 나비":
        await message.channel.send("팔랑팔랑~ :butterfly:")
    if message.content == "치피야 개미":
        await message.channel.send("뚠뚠~ :ant:")
    if message.content == "치피야 울어":
        await message.channel.send("***Miaow!***")
    if message.content == "치피야 어디야":
        await message.channel.send("**여기요!**")
    if message.content.startswith("치피야 그림"):
        await message.channel.send("저는 그림 완전 못그려요.. ㅠㅠ")
    if message.content == "치피야 심심해":
        await message.channel.send("저랑 놀아요!")
    if message.content == "치피야 심심해?":
        await message.channel.send("네! 저랑 놀아주세요!")
    if message.content.startswith("치피야 무슨동물"):
        await message.channel.send("전 고양이랍니다!")
    if message.content.startswith("치피야 코딱지"):
        await message.channel.send("더러워요")
    if message.content.startswith("치피야 념념굿"):
        await message.channel.send("념념굿~")
    if message.content.startswith("치피야 미워"):
        await message.channel.send("힝..*(시무룩)*")
    if message.content.startswith("치피야 처리해"):
        await message.channel.send("네 알겠습니다.")
    if message.content.startswith("치피야 놀자"):
        await message.channel.send("좋아요! 뭐부터 할까요?")
    if message.content.startswith("차피야"):
        await message.channel.send("저는 차피가 아니라구요!")
    if message.content.startswith("차파야"):
        await message.channel.send("저는 차파가 아니고 치피인데요?")
    if message.content.startswith("치피야 백얌"):
        await message.channel.send("백백!!")
    if message.content.startswith("치피야 양초"):
        await message.channel.send("고1이다제~~")

    if message.content.startswith("치피야") and message.author.id != 704319250155700224:
        file = openpyxl.load_workbook("친밀도.xlsx")
        sheet = file.active
        exp = [50, 100, 200, 400, 800, 1000, 3000, 5000, 7500, 10000, 15000]
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = sheet["B" + str(i)].value + 5
                if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value - 1]:
                    sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                    await message.channel.send("[치피와의 친밀도가 올랐습니다!] \n현재 친밀도 :two_hearts: : " + str(sheet["C" + str(i)].value) + "\nEXP : " + str(sheet["B" + str(i)].value))
                file.save("친밀도.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 0
                sheet["C" + str(i)].value = 1
                file.save("친밀도.xlsx")
                break

            i += 1

    if message.content == "치피야 친밀도":
        await message.channel.send("**현재 친밀도 :two_hearts: : **" + str(sheet["C" + str(i)].value) + "**\nEXP : **" + str(sheet["B" + str(i)].value))

    if message.content.startswith("치피야 자기소개"):
        embed = discord.Embed(title="이름 : **치피**", description="소개", color=109243)
        embed.add_field(name="여러분들의 친구! 치피입니다!", value="그 외", inline=False)
        embed.set_footer(text="더이상 설명은 불가능 합니다. 으 귀여워 ㅠㅠ")
        await message.channel.send(embed=embed)

    if message.content.startswith("치피야 내정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=109243)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=False)
        embed.add_field(name="가입일", value=str(date.year) + "년 " + str(date.month) + "월 " + str(date.day) + "일", inline=False)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)

    if message.content.startswith("치피야 도움말"):
        embed = discord.Embed(title="***치피 도움말 입니다!***", description="**치피 명령어**", color=109243)
        embed.add_field(name="**[치피야 도움말]\n[치피야 친밀도]\n[치피야 내정보]\n[!타이머 (초)]\n[치피!배워 (우리가 할말) (치피가 할말)]**", value="※치피를 가르칠때 (우리가 할말)과 (치피가 할말) 에는 띄어쓰기를 하시면 버그가 날수도 있어요!※\n ver.1.6", inline=False)
        await message.channel.send(embed=embed)

    if message.content.startswith("치피!배워"):
        file = openpyxl.load_workbook("치피가 배운것.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 999):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("**네! 잘 배웠어요!**")
                break
        file.save("치피가 배운것.xlsx")

    if message.content.startswith("치피야"):
        file = openpyxl.load_workbook("치피가 배운것.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 999):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith('!타이머'):

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)
        vrsize = int(vrsize)
        for i in range(1, vrsize):
            Text = Text + " " + learn[i]

        sec = int(Text)

        for i in range(sec, 0, -1):
            print(i)
            msg = await message.channel.send(':zap:[타이머 작동중 : '+ str(i) + '초]:zap:')
            time.sleep(1)
            await msg.edit(content=':zap:[타이머 작동중 : '+ str(i) + '초]:zap:')
        else:
            print("땡")
            await message.channel.send(f'{message.author.mention}님! 타이머가 끝났어요! :wink:')







client.run(token)
